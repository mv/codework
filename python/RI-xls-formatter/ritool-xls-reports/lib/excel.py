#!/usr/bin/env python3

# sqlite

from openpyxl                 import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles          import PatternFill, Border, Side, Alignment, Protection, Font


# Excel formats

font_1 = Font(name='Arial', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
font_2 = Font(name='Arial', size=10, bold=True,  italic=True,  vertAlign=None, underline='none', strike=False, color='FF000000')
font_h = Font(name='Arial', size=10, bold=True,  italic=True,  vertAlign=None, underline='none', strike=False, color='FF000000')

formats = {
    'A': { 'font': font_1 ,'width': 14.00, 'number_format':           '0'       }, # linkedaccountid
    'B': { 'font': font_1 ,'width': 16.00, 'number_format':           '@'       }, # Billed instance Type
    'C': { 'font': font_1 ,'width': 10.00, 'number_format':           '@'       }, # OS Type
    'D': { 'font': font_1 ,'width':  9.00, 'number_format':           '@'       }, # region
    'E': { 'font': font_1 ,'width':  8.00, 'number_format':           '@'       }, # tenancy
    'F': { 'font': font_1 ,'width': 14.00, 'number_format':      '##,##0.00000' }, # OnDemand Rate
    'G': { 'font': font_2 ,'width': 10.00, 'number_format':      '##,##0.0'     }, # billed_hours
    'H': { 'font': font_1 ,'width': 21.00, 'number_format': '###,###,##0.00'    }, # Current OnDemand Price
    'I': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.0'     }, # Usage Min
    'J': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.0'     }, # Usage Max
    'K': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.00'    }, # Usage Avg
    'L': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.0'     }, # RI Recommendation
    'M': { 'font': font_2 ,'width': 10.00, 'number_format': '###,###,##0.00'    }, # ri_fee
    'N': { 'font': font_1 ,'width': 13.00, 'number_format': '###,###,##0.00'    }, # RI Upfront Price
    'O': { 'font': font_1 ,'width': 18.00, 'number_format': '###,###,##0.00'    }, # Effective Monthly Price
    'P': { 'font': font_1 ,'width': 19.00, 'number_format': '###,###,##0.00'    }, # Effective Monthly Saving
    'Q': { 'font': font_2 ,'width': 10.00, 'number_format':         '##0.00%'   }, # saving_perc
    'R': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.00'    }, # Break Even
    'S': { 'font': font_1 ,'width': 10.00, 'number_format':      '##,##0.00'    }, # Confidence
    'T': { 'font': font_2 ,'width': 10.00, 'number_format':      '##,##0.0'     }, # monthly_hours
    'U': { 'font': font_1 ,'width': 21.00, 'number_format': '###,###,##0.00'    }, # OnDemand Monthly Price
    'V': { 'font': font_1 ,'width': 16.00, 'number_format': '###,###,##0.00'    }, # Actual Monthly Price
    'W': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.00000' }, # RI Actual Rate
    'X': { 'font': font_1 ,'width': 13.00, 'number_format':      '##,##0.00000' }, # RI Effective Rate
    'Y': { 'font': font_1 ,'width':  8.00, 'number_format':         '##0.00'    }, # ROI %
    'Z': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.00'    }, # Fully Paid Day
}

def create_file(df, file_name):

    # fix
    df['linked_account'] = df['linked_account'].astype(int)

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Format Values
    for col in formats.keys():
        for cell in ws[col]:
            cell.number_format = formats[col]['number_format']
            cell.font          = formats[col]['font']
        ws.column_dimensions[col].width = formats[col]['width']

    # Format Headers
    for cell in ws['1']:
        cell.font = font_h
        cell.fill = PatternFill( fill_type = 'solid', start_color = 'CCCCCC', end_color = 'FFFFFF' )

    wb.save( file_name )


def create_workbook(df_hash, df_totals, file_name):

    wb = Workbook()
    for name in df_hash.keys():
        print('Sheet: ' + name)
        wb.create_sheet(title=name)
        ws = wb.get_sheet_by_name(name)

        df = df_hash[ name ]
        df['linked_account'] = df['linked_account'].astype(int)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Format Values
        for col in formats.keys():
            for cell in ws[col]:
                cell.number_format = formats[col]['number_format']
                cell.font          = formats[col]['font']
            ws.column_dimensions[col].width = formats[col]['width']

        # Format Headers
        for cell in ws['1']:
            cell.font = font_h
            cell.fill = PatternFill( fill_type = 'solid', start_color = 'CCCCCC', end_color = 'FFFFFF' )

    # Final
    ws = wb.get_sheet_by_name('Sheet')
    for name in df_totals.keys():
        print('Totals: ' + name)

#       df = df_totals[ name ]
#       df = df_totals
#       for r in dataframe_to_rows(df, index=False, header=True):
#           ws.append(r)
#
#       # Format Values
#       for col in formats.keys():
#           for cell in ws[col]:
#               cell.number_format = '###,###,##0.00'
#               cell.font          = font_1
#           ws.column_dimensions[col].width = 16.00
#
#       # Format Headers
#       for cell in ws['1']:
#           cell.font = font_h
#           cell.fill = PatternFill( fill_type = 'solid', start_color = 'CCCCCC', end_color = 'FFFFFF' )

    wb.remove_sheet(ws)
    wb.save( file_name )


