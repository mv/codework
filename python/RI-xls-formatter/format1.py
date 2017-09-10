#!/usr/bin/env python3


import sys
import pandas as pd

from openpyxl                 import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles          import PatternFill, Border, Side, Alignment, Protection, Font



# Files in csv format

file = 'DevryB_201708_rgn_fuf_ri_lnk_threeyr.xls'
# file = sys.argv[1]

df = pd.read_csv(file, sep='\t', header=1)
df = df.sort_values(['region', 'OS Type', 'Current OnDemand Price'], ascending=[True, False, False])


# Excel

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

font_1 = Font(name='Arial', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
font_2 = Font(name='Arial', size=10, bold=True,  italic=True,  vertAlign=None, underline='none', strike=False, color='FF000000')
font_h = Font(name='Arial', size=10, bold=True,  italic=True,  vertAlign=None, underline='none', strike=False, color='FF000000')

formats = {
    'A': { 'font': font_1 ,'width': 14.00, 'number_format': '0' },  # linkedaccountid
    'B': { 'font': font_1 ,'width': 16.00, 'number_format': '0' },  # Billed instance Type
    'C': { 'font': font_1 ,'width':  9.00, 'number_format': '0' },  # OS Type
    'D': { 'font': font_1 ,'width': 21.00, 'number_format': '0' },  # region
    'E': { 'font': font_1 ,'width':  8.00, 'number_format': '0' },  # tenancy
    'F': { 'font': font_1 ,'width': 14.00, 'number_format':      '##,##0.00000' },  # OnDemand Rate
    'G': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.00000' },  # RI Actual Rate
    'H': { 'font': font_1 ,'width': 13.00, 'number_format':      '##,##0.00000' },  # RI Effective Rate
    'I': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.0'     },  # Usage Min
    'J': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.0'     },  # Usage Max
    'K': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.00'    },  # Usage Avg
    'L': { 'font': font_1 ,'width': 19.00, 'number_format': '###,###,##0.00000' },  # Current OnDemand Price
    'M': { 'font': font_1 ,'width': 20.00, 'number_format': '###,###,##0.00000' },  # OnDemand Monthly Price
    'N': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.0'     },  # RI Recommendation
    'O': { 'font': font_1 ,'width': 13.00, 'number_format': '###,###,##0.00'    },  # RI Upfront Price
    'P': { 'font': font_1 ,'width': 18.00, 'number_format': '###,###,##0.00000' },  # Effective Monthly Price
    'Q': { 'font': font_1 ,'width': 16.00, 'number_format': '###,###,##0.00000' },  # Actual Monthly Price
    'R': { 'font': font_1 ,'width': 19.00, 'number_format': '###,###,##0.00000' },  # Effective Monthly Saving
    'S': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.00'    },  # Break Even
    'T': { 'font': font_1 ,'width':  8.00, 'number_format':         '##0.00'    },  # ROI %
    'U': { 'font': font_1 ,'width': 11.00, 'number_format':      '##,##0.00'    },  # Fully Paid Day
    'V': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.00'    }   # Confidence
}

# Headers
for cell in ws['1']:
    cell.font = font_h


# Values
for col in formats.keys():
    ws.column_dimensions[col].width = formats[col]['width']
    for cell in ws[col]:
        cell.number_format = formats[col]['number_format']
        cell.font          = formats[col]['font']


# Fixes
ws['D1'] = 'Region'

wb.save("example.xlsx")

