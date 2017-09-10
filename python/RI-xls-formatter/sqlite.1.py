#!/usr/bin/env python3


import sys
import re
import pandas as pd

from openpyxl                 import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles          import PatternFill, Border, Side, Alignment, Protection, Font

import sqlite3
from sqlalchemy import create_engine


###
### Pandas: read original xls(csv) files
###


# file = sys.argv[1]
file = 'DevryB_201708_rgn_fuf_ri_lnk_threeyr.xls'

df = pd.read_csv(file, sep='\t', header=1)

#
# fix column names
#
df = df.rename(columns = {
    'linkedaccountid'          : 'linked_account'           ,
    'Billed Instance Type'     : 'instance_type'            ,
    'OS Type'                  : 'os_type'                  ,
    'region'                   : 'region'                   ,
    'Tenancy'                  : 'tenancy'                  ,
    'OnDemand Rate'            : 'on_demand_rate'           ,
    'RI Actual Rate'           : 'ri_actual_rate'           ,
    'RI Effective Rate'        : 'ri_effective_rate'        ,
    'Usage Min'                : 'usage_min'                ,
    'Usage Max'                : 'usage_max'                ,
    'Usage Avg'                : 'usage_avg'                ,
    'Current OnDemand Price'   : 'current_on_demand_price'  ,
    'OnDemand Monthly Price'   : 'on_demand_monthly_price'  ,
    'RI Recommendation'        : 'ri_recommend'             ,
    'RI Upfront Price'         : 'ri_upfront_price'         ,
    'Effective Monthly Price'  : 'effective_monthly_price'  ,
    'Actual Monthly Price'     : 'actual_monthly_price'     ,
    'Effective Monthly Saving' : 'effective_monthly_saving' ,
    'Break Even'               : 'break_even'               ,
    'ROI %'                    : 'roi_perc'                 ,
    'Fully Paid Day'           : 'fully_paid_day'           ,
    'Confidence'               : 'confidence'               ,
})

#
# fix data types
#
df['linked_account'] = df['linked_account'].astype(str)

#
# fix values
#
re_region  = re.compile( r".*\((.*)\)" )
re_os_type = re.compile( r"-License included|-No license required" )
for index, row in df.iterrows():
    df.loc[ index, 'region' ]  = re_region.sub( r'\1', df.loc[ index, 'region'] )
    df.loc[ index, 'os_type' ] = re_os_type.sub( ''  , df.loc[ index, 'os_type'] )

#
# re-ordering
#
df = df.sort_values(['region', 'os_type', 'current_on_demand_price'], ascending=[True, False, False])

#
# add new calculated values
#
df['billed_hours'] = df['current_on_demand_price']  / df['on_demand_rate']
df['saving_perc']  = df['effective_monthly_saving'] / df['current_on_demand_price']
df['ri_fee']       = df['ri_upfront_price']         / df['ri_recommend']


# sqlite3 ftw!

db = create_engine('sqlite:///temp.db')
#df.to_sql('mydf', db)

# Excel

wb = Workbook()
ws = wb.active

df['linked_account'] = df['linked_account'].astype(int)
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

font_1 = Font(name='Arial', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
font_2 = Font(name='Arial', size=10, bold=True,  italic=True,  vertAlign=None, underline='none', strike=False, color='FF000000')
font_h = Font(name='Arial', size=10, bold=True,  italic=True,  vertAlign=None, underline='none', strike=False, color='FF000000')

formats = {
    'A': { 'font': font_1 ,'width': 14.00, 'number_format':           '0'       }, # linkedaccountid
    'B': { 'font': font_1 ,'width': 16.00, 'number_format':           '@'       }, # Billed instance Type
    'C': { 'font': font_1 ,'width': 10.00, 'number_format':           '@'       }, # OS Type
    'D': { 'font': font_1 ,'width':  9.00, 'number_format':           '@'       }, # region
    'E': { 'font': font_1 ,'width':  8.00, 'number_format':           '@'       }, # tenancy
    'F': { 'font': font_1 ,'width': 14.00, 'number_format':      '##,##0.00000' }, # OnDemand Rate
    'G': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.00000' }, # RI Actual Rate
    'H': { 'font': font_1 ,'width': 13.00, 'number_format':      '##,##0.00000' }, # RI Effective Rate
    'I': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.0'     }, # Usage Min
    'J': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.0'     }, # Usage Max
    'K': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.00'    }, # Usage Avg
    'L': { 'font': font_1 ,'width': 21.00, 'number_format': '###,###,##0.00'    }, # Current OnDemand Price
    'M': { 'font': font_1 ,'width': 21.00, 'number_format': '###,###,##0.00'    }, # OnDemand Monthly Price
    'N': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.0'     }, # RI Recommendation
    'O': { 'font': font_1 ,'width': 13.00, 'number_format': '###,###,##0.00'    }, # RI Upfront Price
    'P': { 'font': font_1 ,'width': 18.00, 'number_format': '###,###,##0.00'    }, # Effective Monthly Price
    'Q': { 'font': font_1 ,'width': 16.00, 'number_format': '###,###,##0.00'    }, # Actual Monthly Price
    'R': { 'font': font_1 ,'width': 19.00, 'number_format': '###,###,##0.00'    }, # Effective Monthly Saving
    'S': { 'font': font_1 ,'width':  9.00, 'number_format':      '##,##0.00'    }, # Break Even
    'T': { 'font': font_1 ,'width':  8.00, 'number_format':         '##0.00'    }, # ROI %
    'U': { 'font': font_1 ,'width': 12.00, 'number_format':      '##,##0.00'    }, # Fully Paid Day
    'V': { 'font': font_1 ,'width': 10.00, 'number_format':      '##,##0.00'    }, # Confidence
    'W': { 'font': font_2 ,'width': 10.00, 'number_format':      '##,##0.0'     }, # billed_hours
    'X': { 'font': font_2 ,'width': 10.00, 'number_format':      '##,##0.00'    }, # saving_perc
    'Y': { 'font': font_2 ,'width': 10.00, 'number_format': '###,###,##0.00'    }, # ri_fee
}

# Values
for col in formats.keys():
    for cell in ws[col]:
        cell.number_format = formats[col]['number_format']
        cell.font          = formats[col]['font']

    ws.column_dimensions[col].width = formats[col]['width']


# Headers
for cell in ws['1']:
    cell.font = font_h
    cell.fill = PatternFill( fill_type = 'solid', start_color = 'CCCCCC', end_color = 'FFFFFF' )


wb.save("example.xlsx")

